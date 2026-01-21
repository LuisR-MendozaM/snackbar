
import os
from datetime import datetime
from openpyxl import load_workbook

class ExcelUnicoArchivo:
    def __init__(self):
        """Inicializa la clase cargando el archivo una vez en memoria"""
        # Ruta del archivo principal
        carpeta_datos="C:/Users/luism/Desktop/Datos_Sistema_Monitoreo"
        self.carpeta_datos = carpeta_datos
        self.archivo = os.path.join(self.carpeta_datos, "Prueba.xlsm")
        
        print(f"Usando: {self.archivo}")

        self.bandera_archivo = True

        # self.archivo = os.path.join(carpeta_datos, "Prueba.xlsm")
        # print(f"📝 Usando: {os.path.basename(self.archivo)}")
        
        if not os.path.exists(self.archivo):
            self.bandera_archivo = False
            # raise FileNotFoundError(f"No se encuentra: {self.archivo}")
        
        # Parámetros para las hojas (AJUSTA los nombres si son diferentes)
        self.parametros = {
            'temperatura': {'nombre': 'Temperatura', 'unidad': '°C'},
            'humedad': {'nombre': 'Humedad', 'unidad': '%'},
            'presion1': {'nombre': 'Presion1', 'unidad': 'Pa'},
            'presion2': {'nombre': 'Presion2', 'unidad': 'Pa'},
            'presion3': {'nombre': 'Presion3', 'unidad': 'Pa'},
        }
        
        # Mapeo de parámetros a hojas
        self.hojas = {
            'temperatura': 'Temperatura',
            'humedad': 'Humedad', 
            'presion1': 'Presion1',
            'presion2': 'Presion2',
            'presion3': 'Presion3',
        }
        
        # Posiciones por mes (AJUSTA ESTO según tu plantilla)
        self.posiciones_mes = {
            1: (3, 32),    # Enero
            2: (37, 66),   # Febrero
            3: (70, 99),   # Marzo
            4: (103, 132), # Abril
            5: (136, 165), # Mayo
            6: (169, 198), # Junio
            7: (202, 231), # Julio
            8: (235, 264), # Agosto
            9: (268, 297), # Septiembre
            10: (301, 330), # Octubre
            11: (334, 363), # Noviembre
            12: (367, 396)  # Diciembre
        }
        
        # Cargar el libro UNA VEZ al iniciar
        try:
            self.wb = load_workbook(self.archivo, keep_vba=True, data_only=False)
            print(f"📘 Libro cargado en memoria: {os.path.basename(self.archivo)}")
        except Exception as e:
            print(f"Error al cargar el archivo: {e}")
    
    def _buscar_fila_vacia(self, ws, mes):
        """Busca la primera fila vacía en el rango del mes especificado"""
        if mes not in self.posiciones_mes:
            return None
            
        inicio, fin = self.posiciones_mes[mes]
        
        # Buscar fila vacía en la columna A
        for fila in range(inicio, fin + 1):
            if ws[f'A{fila}'].value is None:
                return fila
        
        return None
    
    def guardar_dato(self, parametro, valor):
        """Guarda un dato en la plantilla .xlsm (solo en memoria)"""
        try:
            if parametro not in self.hojas:
                print(f"❌ Parámetro no válido: {parametro}")
                return False
            
            # Acceder a la hoja ya cargada en memoria
            ws = self.wb[self.hojas[parametro]]
            
            # Obtener fecha actual
            fecha_completa = datetime.now()
            fecha_solo = fecha_completa.date()  # Solo la parte de la fecha
            mes = fecha_completa.month
            
            # Buscar fila vacía para este mes
            fila = self._buscar_fila_vacia(ws, mes)
            
            if fila is None:
                print(f"⚠️ Tabla llena para mes {mes} en hoja {self.hojas[parametro]}")
                return False
            
            # Escribir fecha (como datetime)
            ws[f'A{fila}'].value = fecha_solo
            ws[f'A{fila}'].number_format = 'DD-MMM-YYYY'
            
            # Escribir valor en columna D (como número)
            ws[f'D{fila}'].value = valor
            
            print(f"✓ {self.hojas[parametro]}: {valor} en fila {fila} (en memoria)")
            return True
            
        except KeyError as e:
            print(f"❌ Hoja no encontrada: {e}")
            return False
        except Exception as e:
            print(f"❌ Error guardando {parametro}: {e}")
            return False
    
    def guardar_todos(self, datos):
        """Guarda todos los parámetros y luego guarda el archivo"""
        print(f"\n💾 Escribiendo datos en memoria...")
        
        # Escribir todos los datos en memoria
        resultados = []
        for parametro, valor in datos.items():
            resultado = self.guardar_dato(parametro, valor)
            resultados.append(resultado)
        
        # Guardar el archivo UNA VEZ al final
        try:
            self.wb.save(self.archivo)
            print(f"✅ Todos los datos guardados físicamente en {os.path.basename(self.archivo)}")
        except Exception as e:
            print(f"❌ Error al guardar el archivo: {e}")
            return False
        
        return all(resultados)
    
    def guardar_y_cerrar(self):
        """Guarda y cierra el libro explícitamente"""
        if hasattr(self, 'wb'):
            try:
                self.wb.save(self.archivo)
                self.wb.close()
                print(f"💾 Libro guardado y cerrado: {os.path.basename(self.archivo)}")
                return True
            except Exception as e:
                print(f"❌ Error al guardar/cerrar: {e}")
                return False
        return False
    
    def __del__(self):
        """Asegura que el libro se cierre si se destruye la instancia"""
        if hasattr(self, 'wb'):
            try:
                self.wb.close()
            except:
                pass

    def get_bandera_archivo(self):
        """Retorna la bandera para que main.py la pueda leer"""
        return self.bandera_archivo

# ===================== EJEMPLO DE USO =====================
if __name__ == "__main__":
    # Ejemplo de cómo usar la clase optimizada
    
    # 1. Crear instancia (carga el archivo una vez)
    try:
        excel = ExcelUnicoArchivo("./datos")  # Cambia "./datos" por tu ruta real
    except Exception as e:
        print(f"Error inicial: {e}")
        exit(1)
    
    # 2. Preparar datos para guardar (simulando 30 parámetros)
    datos_a_guardar = {
        'temperatura': 25.5,
        'humedad': 65.2,
        'presion': 1013.25,
        'frecuencia': 50.0
        # Agrega aquí tus otros 26 parámetros...
    }
    
    # 3. Guardar todos los datos de una vez
    try:
        # Opción 1: Usar guardar_todos (recomendado para múltiples datos)
        exito = excel.guardar_todos(datos_a_guardar)
        
        # Opción 2: Usar guardar_dato individualmente y luego guardar_y_cerrar
        # excel.guardar_dato('temperatura', 25.5)
        # excel.guardar_dato('humedad', 65.2)
        # excel.guardar_y_cerrar()
        
        if exito:
            print("🎉 Operación completada exitosamente")
        else:
            print("⚠️ Hubo algunos errores durante el proceso")
            
    except Exception as e:
        print(f"❌ Error durante el guardado: {e}")
    finally:
        # Asegurar que el libro se cierra
        if 'excel' in locals():
            excel.guardar_y_cerrar()


